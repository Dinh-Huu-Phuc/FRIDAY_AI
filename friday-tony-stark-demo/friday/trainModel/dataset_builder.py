from __future__ import annotations

import json
import random
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from ..core.rag import HashEmbeddingModel, InMemoryVectorStore, RagIndexer, TextChunker, ingest_documents_from_paths
from .config import TrainModelConfig
from .schemas import NormalizedDatasetRow, ScoredSample, TrainingExample


@dataclass(slots=True)
class DatasetBuildResult:
    curated_candidates_path: str
    train_path: str
    valid_path: str
    test_path: str
    train_size: int
    valid_size: int
    test_size: int
    dropped_count: int
    kept_count: int
    manifest_path: str

    def to_dict(self) -> dict[str, object]:
        return {
            "curated_candidates_path": self.curated_candidates_path,
            "train_path": self.train_path,
            "valid_path": self.valid_path,
            "test_path": self.test_path,
            "train_size": self.train_size,
            "valid_size": self.valid_size,
            "test_size": self.test_size,
            "dropped_count": self.dropped_count,
            "kept_count": self.kept_count,
            "manifest_path": self.manifest_path,
        }


class DatasetBuilder:
    """
    Build datasets from:
    - scored runtime/training samples
    - merged JSON/JSONL + XLSX sources in trainModel/data/raw/*
    """

    def __init__(self, config: TrainModelConfig) -> None:
        self.config = config

    def build(self, scored_samples: list[ScoredSample]) -> DatasetBuildResult:
        kept_samples = [item for item in scored_samples if item.keep]
        dropped_count = len(scored_samples) - len(kept_samples)
        examples = [self._to_training_example(item) for item in kept_samples]

        randomizer = random.Random(self.config.random_seed)
        randomizer.shuffle(examples)
        train_examples, valid_examples, test_examples = self._split_examples(examples)

        self.config.ensure_directories()
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        curated_candidates_path = self.config.curated_dir / f"curated_candidates_{stamp}.jsonl"
        self._write_jsonl_examples(curated_candidates_path, examples)

        dataset_dir = self.config.datasets_dir / f"dataset_build_{stamp}"
        dataset_dir.mkdir(parents=True, exist_ok=True)
        train_path = dataset_dir / "train.jsonl"
        valid_path = dataset_dir / "valid.jsonl"
        test_path = dataset_dir / "test.jsonl"
        manifest_path = dataset_dir / "manifest.json"

        self._write_jsonl_examples(train_path, train_examples)
        self._write_jsonl_examples(valid_path, valid_examples)
        self._write_jsonl_examples(test_path, test_examples)
        manifest_payload = {
            "build_time": datetime.now(timezone.utc).isoformat(),
            "kept_count": len(kept_samples),
            "dropped_count": dropped_count,
            "train_size": len(train_examples),
            "valid_size": len(valid_examples),
            "test_size": len(test_examples),
            "train_split_ratio": self.config.train_split_ratio,
            "valid_split_ratio": self.config.valid_split_ratio,
            "test_split_ratio": self.config.test_split_ratio,
            "random_seed": self.config.random_seed,
        }
        manifest_path.write_text(json.dumps(manifest_payload, indent=2, ensure_ascii=False), encoding="utf-8")
        return DatasetBuildResult(
            curated_candidates_path=str(curated_candidates_path),
            train_path=str(train_path),
            valid_path=str(valid_path),
            test_path=str(test_path),
            train_size=len(train_examples),
            valid_size=len(valid_examples),
            test_size=len(test_examples),
            dropped_count=dropped_count,
            kept_count=len(kept_samples),
            manifest_path=str(manifest_path),
        )

    def build_from_sources(
        self,
        *,
        export_xlsx: bool = False,
    ) -> dict[str, Any]:
        """
        Merge JSON/JSONL + XLSX into normalized schema, clean invalid rows,
        split train/valid/test, and export reports/artifacts.
        """

        self.config.ensure_directories()
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        json_rows = self._load_json_rows(self.config.data_raw_json_dir)
        xlsx_rows = self._load_xlsx_rows(self.config.data_raw_xlsx_dir)
        merged_rows, validation = self._merge_and_validate(json_rows=json_rows, xlsx_rows=xlsx_rows)

        randomizer = random.Random(self.config.random_seed)
        randomizer.shuffle(merged_rows)
        train_rows, valid_rows, test_rows = self._split_examples(merged_rows)

        processed_dir = self.config.data_processed_dir / f"dataset_{stamp}"
        processed_dir.mkdir(parents=True, exist_ok=True)
        train_path = processed_dir / "train.jsonl"
        valid_path = processed_dir / "valid.jsonl"
        test_path = processed_dir / "test.jsonl"
        merged_path = processed_dir / "merged.jsonl"
        manifest_path = processed_dir / "manifest.json"
        validation_path = processed_dir / "validation_report.json"

        self._write_jsonl_rows(train_path, train_rows)
        self._write_jsonl_rows(valid_path, valid_rows)
        self._write_jsonl_rows(test_path, test_rows)
        self._write_jsonl_rows(merged_path, merged_rows)

        xlsx_output_path = None
        if export_xlsx:
            xlsx_output_path = processed_dir / "merged.xlsx"
            self._write_xlsx_rows(xlsx_output_path, merged_rows)

        manifest_payload = {
            "created_at": datetime.now(timezone.utc).isoformat(),
            "total_rows": len(merged_rows),
            "train_size": len(train_rows),
            "valid_size": len(valid_rows),
            "test_size": len(test_rows),
            "json_rows": len(json_rows),
            "xlsx_rows": len(xlsx_rows),
            "train_ratio": self.config.train_split_ratio,
            "valid_ratio": self.config.valid_split_ratio,
            "test_ratio": self.config.test_split_ratio,
        }
        manifest_path.write_text(json.dumps(manifest_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        validation_path.write_text(json.dumps(validation, ensure_ascii=False, indent=2), encoding="utf-8")

        rag_index_path = self.build_rag_index_from_knowledge()
        return {
            "merged_path": str(merged_path),
            "train_path": str(train_path),
            "valid_path": str(valid_path),
            "test_path": str(test_path),
            "manifest_path": str(manifest_path),
            "validation_report_path": str(validation_path),
            "xlsx_output_path": str(xlsx_output_path) if xlsx_output_path else None,
            "rag_index_path": str(rag_index_path),
        }

    def build_rag_index_from_knowledge(self) -> Path:
        knowledge_paths = [
            self.config.knowledge_raw_dir / "docs",
            self.config.knowledge_raw_dir / "memories",
            self.config.knowledge_raw_dir / "logs",
            self.config.knowledge_raw_dir / "notes",
        ]
        documents = ingest_documents_from_paths(knowledge_paths)
        chunker = TextChunker(chunk_size=self.config.rag_chunk_size, chunk_overlap=self.config.rag_chunk_overlap)
        chunks = []
        for document in documents:
            chunks.extend(chunker.chunk(document))
        indexer = RagIndexer(embedding_model=HashEmbeddingModel(), store=InMemoryVectorStore())
        indexer.index_chunks(chunks)
        indexer.save_index(self.config.vectordb_index_path)
        return self.config.vectordb_index_path

    def _load_json_rows(self, base_dir: Path) -> list[NormalizedDatasetRow]:
        rows: list[NormalizedDatasetRow] = []
        for path in base_dir.rglob("*"):
            if not path.is_file():
                continue
            if path.suffix.lower() == ".jsonl":
                for line in path.read_text(encoding="utf-8").splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    rows.append(self._normalize_row(json.loads(line), source=str(path)))
            elif path.suffix.lower() == ".json":
                payload = json.loads(path.read_text(encoding="utf-8"))
                if isinstance(payload, list):
                    for item in payload:
                        if isinstance(item, dict):
                            rows.append(self._normalize_row(item, source=str(path)))
                elif isinstance(payload, dict):
                    rows.append(self._normalize_row(payload, source=str(path)))
        return rows

    def _load_xlsx_rows(self, base_dir: Path) -> list[NormalizedDatasetRow]:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except Exception:
            return []

        rows: list[NormalizedDatasetRow] = []
        for path in base_dir.rglob("*.xlsx"):
            workbook = load_workbook(path, read_only=True)
            for sheet in workbook.worksheets:
                headers: list[str] = []
                for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                    values = list(row)
                    if row_index == 0:
                        headers = [str(v).strip() if v is not None else "" for v in values]
                        continue
                    if not any(v is not None and str(v).strip() for v in values):
                        continue
                    payload = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
                    rows.append(self._normalize_row(payload, source=str(path)))
        return rows

    def _write_xlsx_rows(self, path: Path, rows: list[NormalizedDatasetRow]) -> None:
        try:
            from openpyxl import Workbook  # type: ignore[import-not-found]
        except Exception:
            return
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            "id",
            "user_id",
            "session_id",
            "timestamp",
            "source",
            "input_text",
            "target_text",
            "emotion_labels",
            "emotion_scores",
            "intent",
            "memory_tags",
            "quality_score",
            "is_safe",
            "notes",
        ]
        sheet.append(headers)
        for row in rows:
            payload = row.to_dict()
            sheet.append(
                [
                    payload["id"],
                    payload["user_id"],
                    payload["session_id"],
                    payload["timestamp"],
                    payload["source"],
                    payload["input_text"],
                    payload["target_text"],
                    json.dumps(payload["emotion_labels"], ensure_ascii=False),
                    json.dumps(payload["emotion_scores"], ensure_ascii=False),
                    payload["intent"],
                    json.dumps(payload["memory_tags"], ensure_ascii=False),
                    payload["quality_score"],
                    payload["is_safe"],
                    payload["notes"],
                ]
            )
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

    def _merge_and_validate(
        self,
        *,
        json_rows: list[NormalizedDatasetRow],
        xlsx_rows: list[NormalizedDatasetRow],
    ) -> tuple[list[NormalizedDatasetRow], dict[str, Any]]:
        merged: dict[str, NormalizedDatasetRow] = {}
        dropped: list[dict[str, str]] = []
        for row in json_rows:
            if not self._is_valid_row(row):
                dropped.append({"id": row.id, "reason": "invalid_json_row"})
                continue
            merged[row.id] = row
        for row in xlsx_rows:
            if not self._is_valid_row(row):
                dropped.append({"id": row.id, "reason": "invalid_xlsx_row"})
                continue
            base = merged.get(row.id)
            merged[row.id] = self._merge_with_override(base, row) if base is not None else row
        return list(merged.values()), {
            "dropped_rows": dropped,
            "dropped_count": len(dropped),
            "merged_count": len(merged),
        }

    def _merge_with_override(self, base: NormalizedDatasetRow, override: NormalizedDatasetRow) -> NormalizedDatasetRow:
        editable_override = override.to_dict()
        payload = base.to_dict()
        for field_name in ("target_text", "emotion_labels", "emotion_scores", "intent", "memory_tags", "quality_score", "is_safe", "notes"):
            value = editable_override[field_name]
            if value not in ("", [], {}, None):
                payload[field_name] = value
        payload["source"] = override.source or payload["source"]
        return self._normalize_row(payload, source=payload["source"])

    def _normalize_row(self, payload: dict[str, Any], *, source: str) -> NormalizedDatasetRow:
        row_id = str(payload.get("id") or "").strip()
        emotion_labels = payload.get("emotion_labels", [])
        emotion_scores = payload.get("emotion_scores", {})
        memory_tags = payload.get("memory_tags", [])
        if isinstance(emotion_labels, str):
            emotion_labels = [item.strip() for item in emotion_labels.split(",") if item.strip()]
        if isinstance(memory_tags, str):
            memory_tags = [item.strip() for item in memory_tags.split(",") if item.strip()]
        if isinstance(emotion_scores, str):
            try:
                emotion_scores = json.loads(emotion_scores)
            except json.JSONDecodeError:
                emotion_scores = {}
        return NormalizedDatasetRow(
            id=row_id,
            user_id=str(payload.get("user_id") or "").strip(),
            session_id=str(payload.get("session_id") or "").strip(),
            timestamp=float(payload.get("timestamp") or 0.0),
            source=str(payload.get("source") or source),
            input_text=str(payload.get("input_text") or payload.get("instruction") or "").strip(),
            target_text=str(payload.get("target_text") or payload.get("output") or "").strip(),
            emotion_labels=[str(v).strip() for v in emotion_labels if str(v).strip()],
            emotion_scores={str(k): float(v) for k, v in dict(emotion_scores).items()} if isinstance(emotion_scores, dict) else {},
            intent=str(payload.get("intent") or "").strip(),
            memory_tags=[str(v).strip() for v in memory_tags if str(v).strip()],
            quality_score=float(payload.get("quality_score") or 0.0),
            is_safe=bool(payload.get("is_safe", True)),
            notes=str(payload.get("notes") or "").strip(),
        )

    def _is_valid_row(self, row: NormalizedDatasetRow) -> bool:
        if not row.id:
            return False
        if not row.input_text or not row.target_text:
            return False
        if row.timestamp <= 0:
            return False
        return True

    def _split_examples(
        self,
        examples: list[TrainingExample] | list[NormalizedDatasetRow],
    ) -> tuple[list[Any], list[Any], list[Any]]:
        total = len(examples)
        if total == 0:
            return [], [], []
        valid_size = int(total * self.config.valid_split_ratio)
        test_size = int(total * self.config.test_split_ratio)
        train_size = total - valid_size - test_size
        if total >= 3 and train_size <= 0:
            train_size = 1
            if valid_size > test_size:
                valid_size -= 1
            else:
                test_size -= 1
        if total >= 3 and valid_size == 0:
            valid_size = 1
            if train_size > 1:
                train_size -= 1
            elif test_size > 0:
                test_size -= 1
        if total >= 3 and test_size == 0:
            test_size = 1
            if train_size > 1:
                train_size -= 1
            elif valid_size > 1:
                valid_size -= 1
        return (
            examples[:train_size],
            examples[train_size : train_size + valid_size],
            examples[train_size + valid_size :],
        )

    def _to_training_example(self, scored: ScoredSample) -> TrainingExample:
        sample = scored.sample
        sample.dataset_status = "curated"
        return TrainingExample(
            system=self.config.training_system_prompt,
            instruction=sample.user_message,
            output=sample.assistant_message,
            metadata={
                "id": f"{sample.session_id}:{int(sample.timestamp)}",
                "session_id": sample.session_id,
                "user_id": sample.user_id,
                "timestamp": sample.timestamp,
                "source": sample.source,
                "feedback_score": sample.feedback_score,
                "resolved": sample.resolved,
                "safety_status": sample.safety_status,
                "quality_score": scored.quality_score,
                "dataset_status": sample.dataset_status,
            },
        )

    def _write_jsonl_examples(self, path: Path, examples: list[TrainingExample]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as file_obj:
            for item in examples:
                file_obj.write(json.dumps(item.to_dict(), ensure_ascii=False) + "\n")

    def _write_jsonl_rows(self, path: Path, rows: list[NormalizedDatasetRow]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as file_obj:
            for row in rows:
                file_obj.write(json.dumps(row.to_dict(), ensure_ascii=False) + "\n")
